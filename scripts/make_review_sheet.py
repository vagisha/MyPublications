#!/usr/bin/env python3
"""Build an Excel review spreadsheet from data/publications.json.

The sheet lets you review every fetched paper and flag any that are NOT yours.
Column B ("Not mine?") has a Yes/No dropdown — set it to "Yes" for any paper
that isn't you. Papers we auto-suspect (co-authors don't overlap your usual
collaborators) are pre-highlighted and pre-marked for review.

Usage:
  python -X utf8 scripts/make_review_sheet.py \
      --in data/publications.json --out data/publications_review.xlsx
"""
import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Frequent, identifying collaborators — used only to *suspect* possible
# name-collision papers so they can be highlighted for a closer look.
KNOWN_COLLABORATORS = {
    "maccoss", "maclean", "eckels", "eng", "riffle", "deutsch", "shulman",
    "noble", "pratt", "whiteaker", "schilling", "bereman", "vitek",
    "perez-riverol", "pérez‐riverol", "bandeira", "merrihew", "tsantilas",
}


def suspect(pub):
    """Return a reason string if this paper looks like it may not be the author's."""
    authors_l = " ; ".join(pub["authors"]).lower()
    if "sharma" not in authors_l:
        return "No 'Sharma' among listed authors"
    if not any(c in authors_l for c in KNOWN_COLLABORATORS):
        return "No usual collaborators among co-authors — check this is you"
    return ""


HEADERS = [
    ("Row", 6), ("Not mine?", 11), ("Auto-flag / note", 34), ("Year", 7),
    ("Title", 60), ("Authors", 50), ("Venue", 30), ("Type", 14),
    ("Citations", 11), ("DOI", 26), ("Link", 40),
]


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--in", dest="inp", default="data/publications.json")
    ap.add_argument("--out", dest="out", default="data/publications_review.xlsx")
    args = ap.parse_args()

    pubs = json.load(open(args.inp, encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    suspect_fill = PatternFill("solid", fgColor="FFF2CC")  # pale amber
    wrap = Alignment(vertical="top", wrap_text=True)

    # Header row
    for c, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    # Yes/No dropdown for "Not mine?"
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)

    for i, p in enumerate(pubs, start=1):
        r = i + 1
        reason = suspect(p)
        doi = p.get("doi")
        link = ("https://doi.org/" + doi) if doi else (p.get("landing_page_url") or p.get("oa_url") or "")
        row = [
            i,
            "Yes" if reason else "No",
            reason,
            p.get("year"),
            p.get("title"),
            "; ".join(p.get("authors", [])),
            p.get("venue"),
            p.get("type"),
            p.get("cited_by_count"),
            doi or "",
            link,
        ]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap
        dv.add(ws.cell(row=r, column=2))
        if reason:
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = suspect_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(pubs) + 1}"

    # A short instructions note in a second sheet
    info = wb.create_sheet("How to use")
    notes = [
        "How to review this list",
        "",
        "1. Column B 'Not mine?' has a Yes/No dropdown.",
        "   Set it to 'Yes' for any paper that is NOT yours.",
        "2. Rows highlighted amber are ones we auto-suspect (co-authors don't",
        "   match your usual collaborators). They are pre-set to 'Yes' — please",
        "   confirm or change them.",
        "3. Add any comment in column C.",
        "4. Save the file and tell me it's ready; I'll drop the flagged rows.",
        "",
        "Data source: OpenAlex, keyed by ORCID 0000-0003-1922-439X.",
    ]
    for i, line in enumerate(notes, start=1):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 78

    wb.save(args.out)
    n_suspect = sum(1 for p in pubs if suspect(p))
    print(f"Wrote {args.out}: {len(pubs)} papers, {n_suspect} auto-flagged for review.")


if __name__ == "__main__":
    main()
